"""Export Excel des findings — workbook auditeur prêt pour revue manuelle.

Onglets produits :
- Findings : 1 ligne par Finding, avec lien hypertexte interne vers la facture.
- Invoices : dataset complet des factures (référencé par les hyperliens).
- RiskScores : score consolidé 0-100 par facture, trié desc.
- Summary : KPI globaux (n factures, n flaggées, € exposition, par détecteur).

Style : en-têtes colorés (palette identité), bandes alternées, panes figées,
formats monétaires/dates corrects pour Excel FR.
"""

from __future__ import annotations

import io
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

from p2p_fraud.schema import Finding, RiskScore

# Palette identité visuelle du projet
_HEADER_FILL = PatternFill("solid", fgColor="0A1628")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_BAND_FILL = PatternFill("solid", fgColor="F3F4F6")
_CRITICAL_FILL = PatternFill("solid", fgColor="FECACA")
_HIGH_FILL = PatternFill("solid", fgColor="FED7AA")


def _findings_to_dataframe(findings: list[Finding]) -> pd.DataFrame:
    if not findings:
        return pd.DataFrame(
            columns=["invoice_id", "detector", "signal", "severity", "rule_id", "evidence"]
        )
    rows = []
    for f in findings:
        rows.append(
            {
                "invoice_id": f.invoice_id,
                "detector": f.detector,
                "signal": f.signal,
                "severity": f.severity.value,
                "rule_id": f.rule_id,
                "evidence": str(f.evidence),
            }
        )
    return pd.DataFrame(rows)


def _scores_to_dataframe(scores: dict[str, RiskScore]) -> pd.DataFrame:
    if not scores:
        return pd.DataFrame(columns=["invoice_id", "risk_score", "findings_count"])
    rows = [
        {
            "invoice_id": rs.invoice_id,
            "risk_score": rs.score,
            "findings_count": rs.findings_count,
            **{f"score_{k}": v for k, v in rs.breakdown.items()},
        }
        for rs in scores.values()
    ]
    df = pd.DataFrame(rows).fillna(0)
    return df.sort_values("risk_score", ascending=False).reset_index(drop=True)


def _style_header(ws, n_columns: int) -> None:
    for col_idx in range(1, n_columns + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"


def _autofit(ws, df: pd.DataFrame, max_width: int = 60) -> None:
    for col_idx, col in enumerate(df.columns, start=1):
        max_len = max(
            len(str(col)),
            *(len(str(v)) if v is not None else 0 for v in df[col].head(200)),
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, max_width)


def _band_rows(ws, n_rows: int, n_columns: int) -> None:
    for r in range(2, n_rows + 2):
        if r % 2 == 0:
            for c in range(1, n_columns + 1):
                ws.cell(row=r, column=c).fill = _BAND_FILL


def _write_dataframe(ws, df: pd.DataFrame, *, banded: bool = True) -> None:
    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)
    if df.empty:
        return
    _style_header(ws, n_columns=df.shape[1])
    if banded:
        _band_rows(ws, n_rows=df.shape[0], n_columns=df.shape[1])
    _autofit(ws, df)


def _add_invoice_hyperlinks(
    ws, df_findings: pd.DataFrame, invoice_index_map: dict[str, int]
) -> None:
    """Ajoute un lien hypertexte sur chaque `invoice_id` -> ligne de l'onglet Invoices."""
    if df_findings.empty:
        return
    col_idx = list(df_findings.columns).index("invoice_id") + 1
    for r, inv_id in enumerate(df_findings["invoice_id"].astype(str), start=2):
        target_row = invoice_index_map.get(inv_id)
        if target_row is None:
            continue
        cell = ws.cell(row=r, column=col_idx)
        cell.hyperlink = f"#Invoices!A{target_row}"
        cell.font = Font(color="2563EB", underline="single")


def _color_severity(ws, df_findings: pd.DataFrame) -> None:
    if df_findings.empty:
        return
    sev_col_idx = list(df_findings.columns).index("severity") + 1
    for r, sev in enumerate(df_findings["severity"].astype(str), start=2):
        cell = ws.cell(row=r, column=sev_col_idx)
        if sev == "critical":
            cell.fill = _CRITICAL_FILL
            cell.font = Font(bold=True, color="991B1B")
        elif sev == "high":
            cell.fill = _HIGH_FILL


def _write_summary(
    ws, *, df_invoices: pd.DataFrame, df_findings: pd.DataFrame, df_scores: pd.DataFrame
) -> None:
    ws.append(["P2P Fraud Detective FR — Synthèse"])
    ws.append([])
    ws.append(["Indicateur", "Valeur"])
    ws.append(["Factures analysées", len(df_invoices)])
    ws.append(["Findings totaux", len(df_findings)])
    ws.append(
        [
            "Factures flaggées (≥ 1 finding)",
            df_findings["invoice_id"].nunique() if not df_findings.empty else 0,
        ]
    )
    if not df_invoices.empty and "amount" in df_invoices.columns:
        flagged_ids = set(df_findings["invoice_id"].astype(str)) if not df_findings.empty else set()
        exposure = df_invoices.loc[
            df_invoices["invoice_id"].astype(str).isin(flagged_ids), "amount"
        ].sum()
        ws.append(["Exposition € (factures flaggées)", float(exposure)])
    if not df_scores.empty:
        ws.append(["Score risque max", float(df_scores["risk_score"].max())])
        ws.append(["Score risque médian", float(df_scores["risk_score"].median())])
        critique = (df_scores["risk_score"] >= 80).sum()
        eleve = ((df_scores["risk_score"] >= 50) & (df_scores["risk_score"] < 80)).sum()
        ws.append(["Bande CRITIQUE (≥ 80)", int(critique)])
        ws.append(["Bande ÉLEVÉ (50-79)", int(eleve)])

    if not df_findings.empty:
        ws.append([])
        ws.append(["Findings par détecteur", ""])
        for detector, n in df_findings["detector"].value_counts().items():
            ws.append([detector, int(n)])

    # Mise en forme : titre + en-tête tableau
    ws.cell(row=1, column=1).font = Font(bold=True, size=14, color="0A1628")
    ws.merge_cells("A1:B1")
    ws.cell(row=3, column=1).font = _HEADER_FONT
    ws.cell(row=3, column=1).fill = _HEADER_FILL
    ws.cell(row=3, column=2).font = _HEADER_FONT
    ws.cell(row=3, column=2).fill = _HEADER_FILL
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 24


def build_workbook(
    *,
    invoices: pd.DataFrame,
    findings: list[Finding],
    risk_scores: dict[str, RiskScore],
) -> Workbook:
    """Construit le workbook openpyxl complet (en mémoire)."""
    wb = Workbook()
    wb.remove(wb.active)

    # 1) Summary
    ws_summary = wb.create_sheet("Summary")
    df_findings = _findings_to_dataframe(findings)
    df_scores = _scores_to_dataframe(risk_scores)
    _write_summary(ws_summary, df_invoices=invoices, df_findings=df_findings, df_scores=df_scores)

    # 2) RiskScores
    ws_scores = wb.create_sheet("RiskScores")
    _write_dataframe(ws_scores, df_scores)

    # 3) Findings (avec hyperliens vers Invoices)
    ws_findings = wb.create_sheet("Findings")
    _write_dataframe(ws_findings, df_findings)

    # 4) Invoices (référencées par les hyperliens)
    ws_invoices = wb.create_sheet("Invoices")
    _write_dataframe(ws_invoices, invoices)

    invoice_index_map: dict[str, int] = {}
    if not invoices.empty:
        for r, inv_id in enumerate(invoices["invoice_id"].astype(str), start=2):
            invoice_index_map[inv_id] = r

    _add_invoice_hyperlinks(ws_findings, df_findings, invoice_index_map)
    _color_severity(ws_findings, df_findings)

    return wb


def export_findings(
    output: Path | io.BytesIO,
    *,
    invoices: pd.DataFrame,
    findings: list[Finding],
    risk_scores: dict[str, RiskScore],
) -> None:
    """Écrit le workbook sur disque ou dans un buffer (Streamlit download)."""
    wb = build_workbook(invoices=invoices, findings=findings, risk_scores=risk_scores)
    if isinstance(output, Path):
        output.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(output))
    else:
        wb.save(output)
