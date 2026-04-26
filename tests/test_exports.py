"""Tests des exports Excel + Parquet."""

from __future__ import annotations

import io
from datetime import date

import pandas as pd
from openpyxl import load_workbook

from p2p_fraud.export.excel_findings import build_workbook, export_findings
from p2p_fraud.export.parquet_for_powerbi import export_to_parquet
from p2p_fraud.schema import Finding, RiskScore, Severity


def _sample_invoices() -> pd.DataFrame:
    return pd.DataFrame(
        {
            "invoice_id": ["INV1", "INV2", "INV3"],
            "vendor_name": ["X SARL", "Y SAS", "Z SA"],
            "amount": [1000.0, 5000.0, 10000.0],
            "invoice_date": [date(2025, 1, 1), date(2025, 2, 1), date(2025, 3, 1)],
            "siren": ["123456789", "987654321", "555444333"],
            "iban": ["FR111", "FR222", "FR333"],
        }
    )


def _sample_findings() -> list[Finding]:
    return [
        Finding(
            invoice_id="INV1",
            detector="duplicates",
            signal="duplicate_exact",
            severity=Severity.CRITICAL,
            rule_id="DUP_EXACT",
            evidence={"siblings": ["INV99"]},
        ),
        Finding(
            invoice_id="INV2",
            detector="benford",
            signal="benford_anomaly_f2d",
            severity=Severity.MEDIUM,
            rule_id="BENFORD_F2D",
            evidence={"digit_value": 50},
        ),
    ]


def _sample_scores() -> dict[str, RiskScore]:
    return {
        "INV1": RiskScore(
            invoice_id="INV1", score=85.0, findings_count=1, breakdown={"duplicates": 60.0}
        ),
        "INV2": RiskScore(
            invoice_id="INV2", score=18.0, findings_count=1, breakdown={"benford": 18.0}
        ),
    }


def test_excel_workbook_structure():
    wb = build_workbook(
        invoices=_sample_invoices(), findings=_sample_findings(), risk_scores=_sample_scores()
    )
    assert {"Summary", "Findings", "Invoices", "RiskScores"} == set(wb.sheetnames)
    # Findings : 2 lignes + 1 header
    assert wb["Findings"].max_row == 3
    # Invoices : 3 lignes + 1 header
    assert wb["Invoices"].max_row == 4
    # RiskScores : 2 lignes + 1 header
    assert wb["RiskScores"].max_row == 3


def test_excel_findings_have_hyperlinks():
    wb = build_workbook(invoices=_sample_invoices(), findings=_sample_findings(), risk_scores={})
    ws = wb["Findings"]
    invoice_id_col = next(
        c for c in range(1, ws.max_column + 1) if ws.cell(row=1, column=c).value == "invoice_id"
    )
    cell = ws.cell(row=2, column=invoice_id_col)
    assert cell.hyperlink is not None
    assert "Invoices!" in cell.hyperlink.target


def test_excel_round_trip_to_buffer():
    buf = io.BytesIO()
    export_findings(
        buf, invoices=_sample_invoices(), findings=_sample_findings(), risk_scores=_sample_scores()
    )
    buf.seek(0)
    wb = load_workbook(buf)
    assert "Summary" in wb.sheetnames


def test_parquet_export_creates_three_files(tmp_path):
    paths = export_to_parquet(
        tmp_path,
        invoices=_sample_invoices(),
        findings=_sample_findings(),
        risk_scores=_sample_scores(),
    )
    assert set(paths) == {"invoices", "findings", "risk_scores"}
    for p in paths.values():
        assert p.exists()

    invoices_back = pd.read_parquet(paths["invoices"])
    assert len(invoices_back) == 3

    findings_back = pd.read_parquet(paths["findings"])
    assert len(findings_back) == 2
    assert set(findings_back.columns) >= {
        "invoice_id",
        "detector",
        "signal",
        "severity",
        "rule_id",
        "evidence_json",
    }

    scores_back = pd.read_parquet(paths["risk_scores"])
    assert len(scores_back) == 2


def test_parquet_handles_empty_inputs(tmp_path):
    paths = export_to_parquet(tmp_path, invoices=pd.DataFrame(), findings=[], risk_scores={})
    for p in paths.values():
        assert p.exists()
